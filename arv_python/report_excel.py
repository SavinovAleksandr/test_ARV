# -*- coding: utf-8 -*-
"""
Формирование отчёта «Приложение. Результат в табличном виде.xlsx».
Структура: заголовки A1–F2, затем блоки по режимам (название режима, схемы, строки параметров).
"""
import os
from typing import List, Tuple, Optional

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


# Строка данных: (param_name, degree, damp, eff, note)
ExcelRow = Tuple[str, Optional[float], Optional[str], Optional[str], Optional[str]]
SchemeBlock = Tuple[str, Optional[List[ExcelRow]]]  # (scheme_name, rows or None)
RegimeBlock = Tuple[str, List[SchemeBlock]]  # (regime_name, list of schemes)


def write_result_xlsx(
    regime_blocks: List[RegimeBlock],
    sheet_name: str,
    out_path: str,
) -> None:
    """
    Записать итоговый xlsx.
    regime_blocks: список (regime_name, [(scheme_name, None)] или [(scheme_name, [rows])]).
    """
    if openpyxl is None:
        raise RuntimeError("Установите openpyxl: pip install openpyxl")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    # Стили
    font_header = Font(name="Times New Roman", size=10, bold=True)
    font_italic = Font(name="Times New Roman", size=10, italic=True)
    font_red = Font(name="Times New Roman", size=10, bold=True, color="FF0000")
    fill_header = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    fill_regime = PatternFill(start_color="FAFAD2", end_color="FAFAD2", fill_type="solid")
    fill_error = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    # Заголовки
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 20
    ws.merge_cells("A1:A2")
    ws["A1"] = "Схема сети"
    ws.merge_cells("B1:B2")
    ws["B1"] = "Регистрируемый параметр"
    ws.merge_cells("C1:C2")
    ws["C1"] = "Степень демпфирования"
    ws.merge_cells("D1:E1")
    ws["D1"] = "Результат проверки"
    ws["D2"] = "Демпфирование переходного процесса"
    ws["E2"] = "Эффективность PSS"
    ws.merge_cells("F1:F2")
    ws["F1"] = "Примечание"
    for row in (1, 2):
        for col in range(1, 7):
            c = ws.cell(row=row, column=col)
            c.font = font_header
            c.fill = fill_header
            c.alignment = align_center
    row = 3
    for regime_name, schemes in regime_blocks:
        ws.cell(row=row, column=1, value=regime_name)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        for c in range(1, 7):
            ws.cell(row=row, column=c).font = font_italic
            ws.cell(row=row, column=c).fill = fill_regime
            ws.cell(row=row, column=c).alignment = align_center
        row += 1
        for scheme_name, rows in schemes:
            if rows is None:
                ws.cell(row=row, column=1, value=scheme_name)
                ws.cell(row=row, column=2, value="Режим не балансируется в нормальной схеме" if scheme_name == "Нормальная схема" else "Режим не балансируется в ремонтной схеме")
                ws.cell(row=row, column=2).font = font_red
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
                for c in range(1, 7):
                    ws.cell(row=row, column=c).fill = fill_error
                    ws.cell(row=row, column=c).alignment = align_center
                row += 1
                continue
            start_row = row
            for r in rows:
                ws.cell(row=row, column=1, value=scheme_name)
                ws.cell(row=row, column=2, value=r[0])
                ws.cell(row=row, column=3, value=r[1])
                if r[1] is not None:
                    ws.cell(row=row, column=3).number_format = "0.00000"
                ws.cell(row=row, column=4, value=r[2])
                ws.cell(row=row, column=5, value=r[3])
                ws.cell(row=row, column=6, value=r[4] if len(r) > 4 else None)
                if r[2] == "Не демпфируется" or r[3] == "Не эффективен":
                    ws.cell(row=row, column=4).font = font_red
                    ws.cell(row=row, column=5).font = font_red
                for c in range(1, 7):
                    ws.cell(row=row, column=c).alignment = align_center
                row += 1
            if row > start_row:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=row - 1, end_column=1)
    last_row = max(row - 1, 2)
    for r in range(1, last_row + 1):
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = border
    wb.save(out_path)


def build_excel_rows_from_quality(
    quality: List[Tuple[str, float, str, str, Optional[str]]],
) -> List[ExcelRow]:
    """Преобразовать результат get_pss_quality в список строк для Excel: (param, degree, damp, eff, note)."""
    return [(name, degree, damp, eff, note) for name, degree, damp, eff, note in quality]
