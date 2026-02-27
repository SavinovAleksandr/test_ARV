# -*- coding: utf-8 -*-
"""Разбор файла задания xlsx (листы rgms, rems, ut, arv, grf). Соответствует логике compModel.xlxs в C#."""
import os
from pathlib import Path
from typing import List, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from models import (
    ComInfo,
    ElId,
    GrfInfo,
    PssInfo,
    RemsInfo,
    RgmsInfo,
    UtInfo,
)


def _cell(ws: Worksheet, row: int, col: int):
    """Значение ячейки (row/col 1-based)."""
    return ws.cell(row=row, column=col).value


def _str(v) -> str:
    return str(v).strip() if v is not None else ""


def _int(v, default: int = 0) -> int:
    if v is None:
        return default
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return default


def _float(v, default: float = 0.0) -> float:
    if v is None:
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def parse_rgms(ws: Worksheet, regims: List[str]) -> List[RgmsInfo]:
    """Лист rgms: колонки 1=num, 2=name, 3=rId (через ';'). sName подставляется по name из regims."""
    result = []
    row = 2
    while _cell(ws, row, 1) is not None:
        r_id_str = _cell(ws, row, 3)
        if r_id_str is not None:
            r_id = [_int(x) for x in str(r_id_str).split(";") if str(x).strip()]
        else:
            r_id = [0]
        name = _str(_cell(ws, row, 2))
        s_name = None
        for path in regims:
            if Path(path).stem == name:
                s_name = path
                break
        result.append(RgmsInfo(
            num=_int(_cell(ws, row, 1)),
            name=name,
            r_id=r_id,
            s_name=s_name,
        ))
        row += 1
    return result


def parse_rems(ws: Worksheet) -> List[RemsInfo]:
    """Лист rems: строки с 2, колонки 1=id, 2=name, 3,4,5=элемент; следующие строки той же группы если col1 пусто, col3 не пусто. 6=mpdValue, 7=schId, 8=utId."""
    result = []
    row = 2
    while _cell(ws, row, 3) is not None:
        if _cell(ws, row, 1) is None:
            row += 1
            continue
        elements = []
        elements.append(ElId(
            ip=_int(_cell(ws, row, 3)),
            iq=_int(_cell(ws, row, 4)),
            np=_int(_cell(ws, row, 5)),
        ))
        k = row + 1
        while _cell(ws, k, 1) is None and _cell(ws, k, 3) is not None:
            elements.append(ElId(
                ip=_int(_cell(ws, k, 3)),
                iq=_int(_cell(ws, k, 4)),
                np=_int(_cell(ws, k, 5)),
            ))
            k += 1
        result.append(RemsInfo(
            id=_int(_cell(ws, row, 1)),
            name=_str(_cell(ws, row, 2)),
            elements=elements,
            mpd_value=_float(_cell(ws, row, 6), -1.0),
            sch_id=_int(_cell(ws, row, 7), -1),
            ut_id=_int(_cell(ws, row, 8), -1),
        ))
        row += 1
    return result


def parse_ut(ws: Worksheet, ut2_files: List[str]) -> List[UtInfo]:
    """Лист ut: колонки 1=id, 2=name. sName из ut2_files по имени."""
    result = []
    row = 2
    while _cell(ws, row, 1) is not None:
        name = _str(_cell(ws, row, 2))
        s_name = None
        for path in ut2_files:
            if Path(path).stem == name:
                s_name = path
                break
        result.append(UtInfo(
            id=_int(_cell(ws, row, 1)),
            name=name,
            s_name=s_name,
        ))
        row += 1
    return result


def parse_arv(ws: Worksheet) -> List[PssInfo]:
    """Лист arv: колонки 1=table, 2=param, 3=id, 4=val."""
    result = []
    row = 2
    while _cell(ws, row, 1) is not None:
        result.append(PssInfo(
            table=_str(_cell(ws, row, 1)),
            param=_str(_cell(ws, row, 2)),
            id=_int(_cell(ws, row, 3)),
            val=_float(_cell(ws, row, 4)),
        ))
        row += 1
    return result


def parse_grf(ws: Worksheet) -> List[GrfInfo]:
    """Лист grf: колонки 1=table, 2=param, 3=id (через '+'), 4=name."""
    result = []
    row = 2
    while _cell(ws, row, 1) is not None:
        id_str = _str(_cell(ws, row, 3))
        ids = [_int(x) for x in id_str.split("+") if str(x).strip()]
        result.append(GrfInfo(
            table=_str(_cell(ws, row, 1)),
            param=_str(_cell(ws, row, 2)),
            id=ids,
            name=_str(_cell(ws, row, 4)),
        ))
        row += 1
    return result


def load_assignment_xlsx(
    xlsx_path: str,
    regims: List[str],
    ut2_files: List[str],
) -> Optional[ComInfo]:
    """
    Загрузить файл задания xlsx и заполнить ComInfo.
    Возвращает None, если файл не xlsx/xls или нет нужных листов.
    """
    path = Path(xlsx_path)
    if path.suffix.lower() not in (".xlsx", ".xls"):
        return None
    if not path.exists():
        return None
    wb = load_workbook(filename=str(path), read_only=False, data_only=True)
    required = {"rgms", "rems", "ut", "arv"}
    if not required.issubset(wb.sheetnames):
        return None
    info = ComInfo(name=path.stem)
    info.rgms = parse_rgms(wb["rgms"], regims)
    info.rems = parse_rems(wb["rems"])
    info.ut = parse_ut(wb["ut"], ut2_files)
    info.arv = parse_arv(wb["arv"])
    if "grf" in wb.sheetnames:
        info.grf = parse_grf(wb["grf"])
    return info
